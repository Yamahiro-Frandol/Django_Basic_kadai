from django.shortcuts import render, redirect
from django.views import View

from .models import Document

from io import BytesIO
import magic
import openpyxl as px

from pdfminer.high_level import extract_text

ALLOWED_PDF     = [ "application/pdf"]
ALLOWED_EXCEL   = [ "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" ]

# Create your views here.
class IndexView(View):
    def get(self, request, *args, **kwargs):
        # checker/index.html(テンプレート)を表示する。
        context              = {}
        context["documents"] = Document.objects.order_by("-created_at")

        return render(request, "checker/index.html")
        
    def post(self, request, *args, **kwargs):
        
        mime_type   = magic.from_buffer(request.FILES["pdf"].read(1024) , mime=True)
        print(mime_type)
        
        if not mime_type in ALLOWED_PDF:
            print("このファイルはPDFではありません")
        else:

            # 一旦このファイルをメモリ空間内に保存する。
            pdf_virtual = BytesIO(request.FILES["pdf"].read())

            # メモリの読み込み位置を先頭に戻す。
            pdf_virtual.seek(0)

            # メモリの中にあるデータをpdfminerに読ませる。
            text        = extract_text(pdf_virtual)

            # PDFの文字列を読み込みできる。
            print(text)

        mime_type   = magic.from_buffer(request.FILES["excel"].read(1024) , mime=True)
        print(mime_type)
        
        if not mime_type in ALLOWED_EXCEL:
            print("このファイルのEXCELではありません")
        else:

            # TODO: openpyxl は、メモリ空間内に保存されているデータを .load_workbook() で読み込みすることはできない。
            # request.FILES の データをopenpyxlに読ませる。
            wb          = px.load_workbook(request.FILES["excel"])
            ws          = wb.active
            print(ws["A1"].value)
            
            ws["A1"].value  = "test"

        return redirect("checker:index")
    
index   = IndexView.as_view()    